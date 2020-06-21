"""
    是这样的，我的问卷里面有，性别，岗位类别，是否为编制内员工，受教育程度，职称，工作年限，固定休息日，是否遭受过医闹，
    家人对自己的工作是否支持这九个因素，然后满意度有收入是否符合付出，工作是否有自豪感，医患关系处理是否有压力，工作环境是否满意，
    管理体制是否满意，同事相处，发展前景，更换职业倾向这8个，你们能帮我做一个满意度表嘛，就是给8个满意度赋值加权，求平均值，最后出来一个总体满意度

    依赖的包：
        pip install pandas
        pip install xlrd
        pip install openpyxl
"""
import copy
import math
import pandas as pd
import openpyxl


excel_file_name = '80874038_122_122.xlsx'
all_satisfied_data = [[0,0,0] for i in range(8)]


def main():
    def cal_standard_value(values, avearge):
        new_values = [(i-avearge)**2 for i in values]
        return math.sqrt(sum(new_values)/len(values))
    pd_datas = pd.read_excel(excel_file_name, index_col='序号')
    copy1_item = copy.deepcopy(pd_datas.loc[1:, :])
    copy2_item = copy.deepcopy(pd_datas.loc[1:, :])
    pd_datas = pd_datas.append(pd.DataFrame(copy1_item), ignore_index=True)
    pd_datas = pd_datas.append(pd.DataFrame(copy2_item), ignore_index=True)
    pd_datas.index = pd_datas.index + 1
    pd_datas.to_excel("new.{}".format(excel_file_name))
    all_data_len = len(pd_datas)


    wb = openpyxl.Workbook()
    ws = wb.create_sheet(index=0)
    ws2 = wb.create_sheet(index=1)
    pfile = open("results.txt", 'wb')
    pfile2 = open("results_statistic.txt", 'wb')
    item1_key = '9、您认为现在的收入水平符合自己对工作的付出吗'
    item1_key_title = '收入水平符合自己对工作的付出'
    pfile.write("{}\n".format(item1_key_title).encode('utf-8'))
    item1_index2chinese = {
        1: "符合",
        2: "基本符合",
        3: "不符合"
    }
    res = pd_datas.groupby([item1_key])[item1_key].count()
    title_line = '\t\t'
    total_statistic_line = '\t\t'
    total_statistic_persent = '\t\t'
    title_line_list = ['']
    total_statistic_line_list = ['']
    total_statistic_persent_list = ['']
    average_value = 0.0
    for k,v in zip(res.index.tolist(), res.values.tolist()):
        title_line += "{}\t".format(item1_index2chinese[k])
        title_line_list.append(item1_index2chinese[k])
        total_statistic_line += "{}\t".format(v)
        total_statistic_line_list.append(v)
        total_statistic_persent += "{}%\t".format(round(v/all_data_len * 100, 1))
        total_statistic_persent_list.append(round(v/all_data_len * 100, 1))
        average_value += k * (v/all_data_len)
    pfile.write(("{}\n".format(title_line)).encode('utf-8'))
    pfile.write(("{}\n".format(total_statistic_line)).encode('utf-8'))
    pfile.write(("{}\n".format(total_statistic_persent)).encode('utf-8'))
    all_satisfied_data[0][0] = item1_key_title
    all_satisfied_data[0][1] = average_value
    all_satisfied_data[0][2] = cal_standard_value(res.index.tolist(), average_value)
    ws.append([item1_key_title])
    ws.append(title_line_list)
    ws.append(total_statistic_line_list)
    ws.append(total_statistic_persent_list)

    item1_key = '13、您对您的工作有自豪感吗？'
    item1_key_title = '工作是否有自豪感'
    pfile.write("{}\n".format(item1_key_title).encode('utf-8'))
    item1_index2chinese = {
        1: "有",
        2: "没有",
        3: "没有考虑过"
    }
    res = pd_datas.groupby([item1_key])[item1_key].count()
    title_line = '\t\t'
    total_statistic_line = '\t\t'
    total_statistic_persent = '\t\t'
    title_line_list = ['']
    total_statistic_line_list = ['']
    total_statistic_persent_list = ['']
    average_value = 0.0
    for k,v in zip(res.index.tolist(), res.values.tolist()):
        title_line += "{}\t".format(item1_index2chinese[k])
        title_line_list.append(item1_index2chinese[k])
        total_statistic_line += "{}\t".format(v)
        total_statistic_line_list.append(v)
        total_statistic_persent += "{}%\t".format(round(v/all_data_len * 100, 1))
        total_statistic_persent_list.append(round(v/all_data_len * 100, 1))
        average_value += k * (v / all_data_len)
    pfile.write(("{}\n".format(title_line)).encode('utf-8'))
    pfile.write(("{}\n".format(total_statistic_line)).encode('utf-8'))
    pfile.write(("{}\n".format(total_statistic_persent)).encode('utf-8'))
    all_satisfied_data[1][0] = item1_key_title
    all_satisfied_data[1][1] = average_value
    all_satisfied_data[1][2] = cal_standard_value(res.index.tolist(), average_value)
    ws.append([item1_key_title])
    ws.append(title_line_list)
    ws.append(total_statistic_line_list)
    ws.append(total_statistic_persent_list)

    item1_key = '15、您觉得在就诊过程中医患关系的处理对你有压力吗'
    item1_key_title = '医患关系处理是否有压力'
    pfile.write("{}\n".format(item1_key_title).encode('utf-8'))
    item1_index2chinese = {
        1: "压力很大",
        2: "一般",
        3: "没有压力"
    }
    res = pd_datas.groupby([item1_key])[item1_key].count()
    title_line = '\t\t'
    total_statistic_line = '\t\t'
    total_statistic_persent = '\t\t'
    title_line_list = ['']
    total_statistic_line_list = ['']
    total_statistic_persent_list = ['']
    average_value = 0.0
    for k,v in zip(res.index.tolist(), res.values.tolist()):
        title_line += "{}\t".format(item1_index2chinese[k])
        title_line_list.append(item1_index2chinese[k])
        total_statistic_line += "{}\t".format(v)
        total_statistic_line_list.append(v)
        total_statistic_persent += "{}%\t".format(round(v/all_data_len * 100, 1))
        total_statistic_persent_list.append(round(v/all_data_len * 100, 1))
        average_value += k * (v / all_data_len)
    pfile.write(("{}\n".format(title_line)).encode('utf-8'))
    pfile.write(("{}\n".format(total_statistic_line)).encode('utf-8'))
    pfile.write(("{}\n".format(total_statistic_persent)).encode('utf-8'))
    all_satisfied_data[2][0] = item1_key_title
    all_satisfied_data[2][1] = average_value
    all_satisfied_data[2][2] = cal_standard_value(res.index.tolist(), average_value)
    ws.append([item1_key_title])
    ws.append(title_line_list)
    ws.append(total_statistic_line_list)
    ws.append(total_statistic_persent_list)

    item1_key = '17、您对现在的工作环境是否满意'
    item1_key_title = '工作环境是否满意'
    pfile.write("{}\n".format(item1_key_title).encode('utf-8'))
    item1_index2chinese = {
        1: "很满意",
        2: "满意",
        3: "一般",
        4: "不太满意"
    }
    res = pd_datas.groupby([item1_key])[item1_key].count()
    title_line = '\t\t'
    total_statistic_line = '\t\t'
    total_statistic_persent = '\t\t'
    title_line_list = ['']
    total_statistic_line_list = ['']
    total_statistic_persent_list = ['']
    average_value = 0.0
    for k,v in zip(res.index.tolist(), res.values.tolist()):
        title_line += "{}\t".format(item1_index2chinese[k])
        title_line_list.append(item1_index2chinese[k])
        total_statistic_line += "{}\t".format(v)
        total_statistic_line_list.append(v)
        total_statistic_persent += "{}%\t".format(round(v/all_data_len * 100, 1))
        total_statistic_persent_list.append(round(v/all_data_len * 100, 1))
        average_value += k * (v / all_data_len)
    pfile.write(("{}\n".format(title_line)).encode('utf-8'))
    pfile.write(("{}\n".format(total_statistic_line)).encode('utf-8'))
    pfile.write(("{}\n".format(total_statistic_persent)).encode('utf-8'))
    all_satisfied_data[3][0] = item1_key_title
    all_satisfied_data[3][1] = average_value
    all_satisfied_data[3][2] = cal_standard_value(res.index.tolist(), average_value)
    ws.append([item1_key_title])
    ws.append(title_line_list)
    ws.append(total_statistic_line_list)
    ws.append(total_statistic_persent_list)

    item1_key = '18、您认为医院和科室领导的管理体制合理吗'
    item1_key_title = '管理体制是否满意'
    pfile.write("{}\n".format(item1_key_title).encode('utf-8'))
    item1_index2chinese = {
        1: "合理",
        2: "一般",
        3: "不太合理"
    }
    res = pd_datas.groupby([item1_key])[item1_key].count()
    title_line = '\t\t'
    total_statistic_line = '\t\t'
    total_statistic_persent = '\t\t'
    title_line_list = ['']
    total_statistic_line_list = ['']
    total_statistic_persent_list = ['']
    average_value = 0.0
    for k,v in zip(res.index.tolist(), res.values.tolist()):
        title_line += "{}\t".format(item1_index2chinese[k])
        title_line_list.append(item1_index2chinese[k])
        total_statistic_line += "{}\t".format(v)
        total_statistic_line_list.append(v)
        total_statistic_persent += "{}%\t".format(round(v/all_data_len * 100, 1))
        total_statistic_persent_list.append(round(v/all_data_len * 100, 1))
        average_value += k * (v / all_data_len)
    pfile.write(("{}\n".format(title_line)).encode('utf-8'))
    pfile.write(("{}\n".format(total_statistic_line)).encode('utf-8'))
    pfile.write(("{}\n".format(total_statistic_persent)).encode('utf-8'))
    all_satisfied_data[4][0] = item1_key_title
    all_satisfied_data[4][1] = average_value
    all_satisfied_data[4][2] = cal_standard_value(res.index.tolist(), average_value)
    ws.append([item1_key_title])
    ws.append(title_line_list)
    ws.append(total_statistic_line_list)
    ws.append(total_statistic_persent_list)

    item1_key = '19、与同事之间的相处对本职工作有什么影响'
    item1_key_title = '同事相处'
    pfile.write("{}\n".format(item1_key_title).encode('utf-8'))
    item1_index2chinese = {
        1: "正面影响",
        2: "负面影响",
        3: "不在意"
    }
    res = pd_datas.groupby([item1_key])[item1_key].count()
    title_line = '\t\t'
    total_statistic_line = '\t\t'
    total_statistic_persent = '\t\t'
    title_line_list = ['']
    total_statistic_line_list = ['']
    total_statistic_persent_list = ['']
    average_value = 0.0
    for k,v in zip(res.index.tolist(), res.values.tolist()):
        title_line += "{}\t".format(item1_index2chinese[k])
        title_line_list.append(item1_index2chinese[k])
        total_statistic_line += "{}\t".format(v)
        total_statistic_line_list.append(v)
        total_statistic_persent += "{}%\t".format(round(v/all_data_len * 100, 1))
        total_statistic_persent_list.append(round(v/all_data_len * 100, 1))
        average_value += k * (v / all_data_len)
    pfile.write(("{}\n".format(title_line)).encode('utf-8'))
    pfile.write(("{}\n".format(total_statistic_line)).encode('utf-8'))
    pfile.write(("{}\n".format(total_statistic_persent)).encode('utf-8'))
    all_satisfied_data[5][0] = item1_key_title
    all_satisfied_data[5][1] = average_value
    all_satisfied_data[5][2] = cal_standard_value(res.index.tolist(), average_value)
    ws.append([item1_key_title])
    ws.append(title_line_list)
    ws.append(total_statistic_line_list)
    ws.append(total_statistic_persent_list)

    item1_key = '20、您认为现在的工作有发展前景吗'
    item1_key_title = '发展前景'
    pfile.write("{}\n".format(item1_key_title).encode('utf-8'))
    item1_index2chinese = {
        1: "有",
        2: "一般",
        3: "不太有"
    }
    res = pd_datas.groupby([item1_key])[item1_key].count()
    title_line = '\t\t'
    total_statistic_line = '\t\t'
    total_statistic_persent = '\t\t'
    title_line_list = ['']
    total_statistic_line_list = ['']
    total_statistic_persent_list = ['']
    average_value = 0.0
    for k,v in zip(res.index.tolist(), res.values.tolist()):
        title_line += "{}\t".format(item1_index2chinese[k])
        title_line_list.append(item1_index2chinese[k])
        total_statistic_line += "{}\t".format(v)
        total_statistic_line_list.append(v)
        total_statistic_persent += "{}%\t".format(round(v/all_data_len * 100, 1))
        total_statistic_persent_list.append(round(v/all_data_len * 100, 1))
        average_value += k * (v / all_data_len)
    pfile.write(("{}\n".format(title_line)).encode('utf-8'))
    pfile.write(("{}\n".format(total_statistic_line)).encode('utf-8'))
    pfile.write(("{}\n".format(total_statistic_persent)).encode('utf-8'))
    all_satisfied_data[6][0] = item1_key_title
    all_satisfied_data[6][1] = average_value
    all_satisfied_data[6][2] = cal_standard_value(res.index.tolist(), average_value)
    ws.append([item1_key_title])
    ws.append(title_line_list)
    ws.append(total_statistic_line_list)
    ws.append(total_statistic_persent_list)

    item1_key = '21、您想更换职业吗'
    item1_key_title = '更换职业倾向'
    pfile.write("{}\n".format(item1_key_title).encode('utf-8'))
    item1_index2chinese = {
        1: "非常想",
        2: "想",
        3: "在考虑",
        4: "不想",
        5: "很不想"
    }
    res = pd_datas.groupby([item1_key])[item1_key].count()
    title_line = '\t\t'
    total_statistic_line = '\t\t'
    total_statistic_persent = '\t\t'
    title_line_list = ['']
    total_statistic_line_list = ['']
    total_statistic_persent_list = ['']
    average_value = 0.0
    for k,v in zip(res.index.tolist(), res.values.tolist()):
        title_line += "{}\t".format(item1_index2chinese[k])
        title_line_list.append(item1_index2chinese[k])
        total_statistic_line += "{}\t".format(v)
        total_statistic_line_list.append(v)
        total_statistic_persent += "{}%\t".format(round(v/all_data_len * 100, 1))
        total_statistic_persent_list.append(round(v/all_data_len * 100, 1))
        average_value += k * (v / all_data_len)
    pfile.write(("{}\n".format(title_line)).encode('utf-8'))
    pfile.write(("{}\n".format(total_statistic_line)).encode('utf-8'))
    pfile.write(("{}\n".format(total_statistic_persent)).encode('utf-8'))
    all_satisfied_data[7][0] = item1_key_title
    all_satisfied_data[7][1] = average_value
    all_satisfied_data[7][2] = cal_standard_value(res.index.tolist(), average_value)
    ws.append([item1_key_title])
    ws.append(title_line_list)
    ws.append(total_statistic_line_list)
    ws.append(total_statistic_persent_list)

    # wb2 = openpyxl.Workbook()
    # ws2 = wb2.create_sheet(index=0)

    new_all_satisfied_data = sorted(all_satisfied_data, key=lambda x:x[1], reverse=True)
    pfile2.write("指标\t均值\t标准差\t排序\n".encode('utf-8'))
    ws2.append(["指标","均值","标准差","排序"])
    for i in range(len(new_all_satisfied_data)):
        pfile2.write(("{}\t{}\t{}\t{}\n".format(new_all_satisfied_data[i][0],
                                                round(new_all_satisfied_data[i][1], 2),
                                                round(new_all_satisfied_data[i][2], 2),
                                                i+1)).encode('utf-8'))
        ws2.append([new_all_satisfied_data[i][0],
                    round(new_all_satisfied_data[i][1], 2),
                    round(new_all_satisfied_data[i][2], 2),
                    i+1])
    pfile.close()
    pfile2.close()
    wb.save('result.xlsx')
    # wb2.save("result_statistic.xlsx")

    # 克伦巴赫a信度系数
    t_row = pd_datas.sum(axis=1)
    # 总方差
    sy = t_row.var()
    # 方差和
    si = pd_datas.var().sum()
    r = (len(pd_datas)/(len(pd_datas)-1)) * (1 - si/sy)
    print("克伦巴赫a信度为：{}".format(r))


main()






