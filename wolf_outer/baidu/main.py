import requests
import time
import random
from lxml import etree
import re
import openpyxl


str_set = {'辞职',
           '离开',
           '结束',
           '不再担任',
           '卸任',
           '出任'}

search_pattern_list = [
    '离职',
    '出任'
]

f = open("{}_{}.txt".format('results', time.time()), 'wb')

def read_excel(excel_file_name):
    wb = openpyxl.load_workbook(excel_file_name)

    ws = wb.get_sheet_by_name(wb.sheetnames[0])
    lines = []
    for i in range(2, ws.max_row+1):
        line = ''
        for j in range(2, ws.max_column+1):
            line += "{} ".format(ws.cell(i,j).value)
        lines.append(line)

    wb.close()
    return lines


def main(word, time_pre):
    url = 'http://www.baidu.com/s'
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36",
        "X-Requested-With": "XMLHttpRequest"
    }

    print("{} {} ===================== ".format(word, time_pre))
    f.write("{} {} ===================== \n".format(word, time_pre).encode('utf-8'))
    for p in search_pattern_list:
        params = {
            "ie": "utf-8",
            "cl": "3",
            "mod": "1",
            "isbd": "1",
            "isid": "A801978E9BB34868",
            "f": "8",
            "rsv_bp": "1",
            "rsv_idx": "2",
            "tn": "baiduhome_pg",
            "wd": "{} {}".format(word, p),
            "rsv_spt": "1",
            "rsv_pq": "c666451600029c2c",
            "rqlang": "cn",
            "rsv_enter": "0",
            "rsv_dl": "tb",
            "rsv_btype": "t",
            "rsv_sid": "undefined",
            "_ss": "1",
            "clist": "",
            "hsug": "",
            "f4s": "1",
            "csor": "0",
            "_cr1": "39593"
        }

        try:
            res = requests.get(url, headers=headers, params=params)
            html = etree.HTML(res.text)
            for item in html.xpath('//div[@class="result c-container "]'):
                item_str = item.xpath('string(./div[@class="c-abstract"])')
                for set_item in str_set:
                    if re.findall(set_item, item_str):
                        print(item_str)
                        f.write("{}\n".format(item_str).encode('utf-8'))
        except Exception as e:
            print('err {}'.format(e))
        else:
            pass
        finally:
            time.sleep(random.randint(1, 4))


if __name__ == '__main__':
    datas = read_excel('ggg.xlsx')
    for data in datas[:2]:
        data_items = data.split()
        if len(data_items) == 3:
            time_items = data_items[0].split('-')
            main("{} {}".format(data_items[1], data_items[2]), time_items[0])

    f.close()


